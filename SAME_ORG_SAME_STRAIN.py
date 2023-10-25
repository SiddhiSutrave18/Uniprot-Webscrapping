import pandas as pd
from openpyxl.styles import PatternFill

# <------------------------Write Input file name----------------------------------------------------->
df = pd.read_csv('SN_MlaA_with_Strain_modified.csv')
# <-------------------------------------------------------------------------------------------------->

df['count'] = df.groupby(['Organism', 'Strain'])['Organism'].transform('count')

# Filter rows where the count is greater than or equal to 2
filtered_df = df[df['count'] >= 2]

# Drop the 'count' column if you don't need it anymore
filtered_df = filtered_df.drop(columns=['count'])
filtered_df = filtered_df.sort_values(by='Strain')
filtered_df = filtered_df[~filtered_df['Strain'].str.contains('Strain Number not found')]
filtered_df.to_csv('Same_Organism_Same_Strain.csv')

writer = pd.ExcelWriter('colored_2_rows.xlsx', engine='openpyxl')

# Write the filtered DataFrame to the Excel file
filtered_df.to_excel(writer, sheet_name='Sheet1', index=False)

# Get the xlsxwriter workbook and worksheet objects
workbook = writer.book
worksheet = writer.sheets['Sheet1']

# Define a dictionary to map strains to colors
strain_colors = {}
colors = [PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # Red
          PatternFill(start_color='B6D7A8', end_color='B6D7A8', fill_type='solid')]  # Green

# Assign a color to each unique strain
for i, strain in enumerate(filtered_df['Strain'].unique()):
    strain_colors[strain] = colors[i % len(colors)]

# Apply colors to the rows based on the 'Strain' column
for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=len(filtered_df) + 1, max_col=len(filtered_df.columns)):
    strain_value = row[filtered_df.columns.get_loc('Strain')].value
    color = strain_colors.get(strain_value, None)
    if color:
        for cell in row:
            cell.fill = color

# <------------------------Write output file name----------------------------------------------------->
workbook.save('colored_2_rows.xlsx')
# <--------------------------------------------------------------------------------------------------->
